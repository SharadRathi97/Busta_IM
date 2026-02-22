from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import FinishedProduct


styles = getSampleStyleSheet()


def _line_cost(qty_per_unit: Decimal, cost_per_unit: Decimal) -> Decimal:
    return (qty_per_unit * cost_per_unit).quantize(Decimal("0.001"))


def bom_to_excel(product: FinishedProduct) -> bytes:
    items = list(product.bom_items.select_related("material").all())
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    ws["A1"] = f"BOM - {product.name}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Product"
    ws["B3"] = product.name
    ws["A4"] = "SKU"
    ws["B4"] = product.sku

    ws.append([])
    ws.append(["S.No", "Raw Material", "Code", "Qty / Unit", "Unit", "Cost / Unit", "Line Cost"])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    total_cost = Decimal("0.000")
    for index, item in enumerate(items, start=1):
        line_cost = _line_cost(item.qty_per_unit, item.material.cost_per_unit)
        total_cost += line_cost
        ws.append(
            [
                index,
                item.material.name,
                item.material.code,
                float(item.qty_per_unit),
                item.material.unit,
                float(item.material.cost_per_unit),
                float(line_cost),
            ]
        )

    ws.append([])
    ws.append(["", "", "", "", "", "Total BOM Cost", float(total_cost)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def bom_to_pdf(product: FinishedProduct) -> bytes:
    items = list(product.bom_items.select_related("material").all())
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    elements = []

    elements.append(Paragraph(f"BOM - {product.name}", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"SKU: {product.sku}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    data = [["S.No", "Raw Material", "Code", "Qty / Unit", "Unit", "Cost / Unit", "Line Cost"]]
    total_cost = Decimal("0.000")
    for index, item in enumerate(items, start=1):
        line_cost = _line_cost(item.qty_per_unit, item.material.cost_per_unit)
        total_cost += line_cost
        data.append(
            [
                str(index),
                item.material.name,
                item.material.code,
                str(item.qty_per_unit),
                item.material.unit,
                str(item.material.cost_per_unit),
                str(line_cost),
            ]
        )

    table = Table(data, colWidths=[14 * mm, 54 * mm, 24 * mm, 24 * mm, 16 * mm, 22 * mm, 22 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Total BOM Cost: {total_cost}", styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()
