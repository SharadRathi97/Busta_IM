from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import PurchaseOrder


styles = getSampleStyleSheet()
STYLE_NORMAL_SMALL = ParagraphStyle("normal_small", parent=styles["Normal"], fontSize=8, leading=10)
STYLE_BOLD_SMALL = ParagraphStyle("bold_small", parent=styles["Normal"], fontSize=8, leading=10, fontName="Helvetica-Bold")
STYLE_BOLD_MEDIUM = ParagraphStyle("bold_medium", parent=styles["Normal"], fontSize=10, leading=12, fontName="Helvetica-Bold")
STYLE_TABLE_CELL_LEFT = ParagraphStyle(
    "table_cell_left",
    parent=STYLE_NORMAL_SMALL,
    alignment=TA_LEFT,
    wordWrap="CJK",
    splitLongWords=True,
)
STYLE_TABLE_CELL_CENTER = ParagraphStyle(
    "table_cell_center",
    parent=STYLE_NORMAL_SMALL,
    alignment=TA_CENTER,
    wordWrap="CJK",
    splitLongWords=True,
)
STYLE_TABLE_HEADER = ParagraphStyle(
    "table_header",
    parent=STYLE_BOLD_SMALL,
    alignment=TA_CENTER,
    wordWrap="CJK",
    splitLongWords=True,
)

COMPANY_NAME = "BUSTA MOBILITY"
COMPANY_BILL_TO_LINES = [
    "Bill To",
    "BUSTA MOBILITY",
    "Shed C, B 13/11, Industrial Estate",
    "Kichha Bypass Road, Rudrapur / U S Nagar-263153",
    "GSTIN NO - 05ABGFB3436Q1Z6 , PAN NO - ABGFB3436Q",
    "STATE NAME - Uttarakhand , State Code - 05",
    "MOB.NO - +91 6395076234/9401145970",
    "Mail Id - admin@bustamobility.com",
]
JURISDICTION_TEXT = "SUBJECT TO RUDRAPUR JURISDICTION"
COMPUTER_GENERATED_TEXT = "This is a Computer Generated Document"

PROJECT_ROOT = Path(settings.BASE_DIR).parent
PREPARED_SIGNATURE_CANDIDATES = [
    PROJECT_ROOT / "assets" / "images" / "prepared_checked_by.png",
    Path(settings.BASE_DIR) / "static" / "img" / "signatures" / "prepared_checked.png",
]
AUTHORISED_SIGNATURE_CANDIDATES = [
    PROJECT_ROOT / "assets" / "images" / "authorized_signatory.png",
    Path(settings.BASE_DIR) / "static" / "img" / "signatures" / "authorised_signatory.png",
]


def _safe_decimal(value) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value))


def _format_qty(value: Decimal) -> str:
    return f"{_safe_decimal(value):,.3f}"


def _format_money(value: Decimal) -> str:
    return f"{_safe_decimal(value):,.2f}"


def _po_number(po: PurchaseOrder) -> str:
    return f"PO-BM-{po.id:02d}"


def _po_doc_number(po: PurchaseOrder) -> str:
    return f"BM/PU/{po.id:02d}"


def _format_date(value) -> str:
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _po_meta_rows(po: PurchaseOrder) -> list[list[str]]:
    revision_date = po.created_at.date().isoformat() if po.created_at else _format_date(po.order_date)
    return [
        ["P.O. No.", _po_number(po)],
        ["P.O. Date", _format_date(po.order_date)],
        ["P.O. Effective Date", _format_date(po.order_date)],
        ["P.O. Validity End Date", _format_date(po.order_date)],
        ["P.O. Amendment No", "-"],
        ["P.O. Amendment Date", "-"],
        ["Vendor Code", po.vendor.vendor_id or "-"],
        ["DOC", _po_doc_number(po)],
        ["Revision No.", "00"],
        ["Revision Date", revision_date],
        ["Buyer Name", po.created_by.get_full_name() if po.created_by and po.created_by.get_full_name() else (po.created_by.username if po.created_by else "-")],
    ]


def _load_signature_image(path: Path, *, max_width_mm: float, max_height_mm: float):
    max_width = max_width_mm * mm
    max_height = max_height_mm * mm
    if not path.exists():
        return Spacer(1, max_height)
    try:
        source_width, source_height = ImageReader(str(path)).getSize()
        if not source_width or not source_height:
            return Spacer(1, max_height)
        scale = min(max_width / source_width, max_height / source_height)
        return RLImage(str(path), width=source_width * scale, height=source_height * scale)
    except Exception:
        return Spacer(1, max_height)


def _resolve_signature_path(candidates: list[Path]) -> Path:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _pdf_cell(value, *, align: str = "left", bold: bool = False) -> Paragraph:
    style = STYLE_TABLE_CELL_CENTER if align == "center" else STYLE_TABLE_CELL_LEFT
    text = escape(str(value if value is not None else "-"))
    if bold:
        text = f"<b>{text}</b>"
    return Paragraph(text, style)


def purchase_order_to_excel(po: PurchaseOrder) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.merge_cells("A1:H1")
    ws["A1"] = "PURCHASE ORDER"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = COMPANY_NAME
    ws["A2"].font = Font(size=11, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 3
    for line in COMPANY_BILL_TO_LINES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=line)
        row += 1

    meta_row = 2
    for label, value in _po_meta_rows(po):
        ws.cell(row=meta_row, column=5, value=label)
        ws.merge_cells(start_row=meta_row, start_column=6, end_row=meta_row, end_column=8)
        ws.cell(row=meta_row, column=6, value=value)
        meta_row += 1

    vendor_start = max(row, meta_row) + 1
    ws.merge_cells(start_row=vendor_start, start_column=1, end_row=vendor_start, end_column=8)
    ws.cell(row=vendor_start, column=1, value="Vendor's Name & Address").font = Font(bold=True)

    vendor_lines = [
        po.vendor.name,
        po.vendor.address_line1,
        po.vendor.address_line2,
        f"{po.vendor.city}, {po.vendor.state}",
        f"State Code - {po.vendor.gst_number[:2] if po.vendor.gst_number else '-'}",
        f"GSTIN/UIN - {po.vendor.gst_number}",
    ]
    data_row = vendor_start + 1
    for line in vendor_lines:
        if not line:
            continue
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=8)
        ws.cell(row=data_row, column=1, value=line)
        data_row += 1

    data_row += 1
    headers = ["Sr.No", "Item Code", "Item Name", "HSN/SAC No", "Qty.", "UOM", "Rate", "Amount"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=data_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cell.border = border

    total_amount = Decimal("0")
    for index, item in enumerate(po.items.select_related("material").all(), start=1):
        data_row += 1
        rate = item.unit_rate if item.unit_rate > 0 else item.material.cost_per_unit
        amount = item.line_amount if item.line_amount > 0 else (item.quantity * rate)
        total_amount += amount

        row_data = [
            index,
            item.material.rm_id or item.material.code,
            item.material.name,
            item.material.code or "-",
            _format_qty(item.quantity),
            item.unit,
            _format_money(rate),
            _format_money(amount),
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=data_row, column=col, value=value)
            cell.border = border
            if col in {1, 4, 5, 6, 7, 8}:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    data_row += 1
    ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=7)
    ws.cell(row=data_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=data_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=data_row, column=8, value=_format_money(total_amount)).font = Font(bold=True)
    for col in range(1, 9):
        ws.cell(row=data_row, column=col).border = border
    ws.cell(row=data_row, column=8).alignment = Alignment(horizontal="center")

    data_row += 2
    ws.cell(row=data_row, column=1, value="TERMS & CONDITIONS").font = Font(bold=True)
    ws.cell(row=data_row, column=2, value=f"GST {po.get_freight_terms_display().upper()}").font = Font(bold=True)
    data_row += 1
    terms_rows = [
        ("1. PAYMENTS", f"PDC against {po.payment_pdc_days} Days"),
        ("2. DELIVERY", po.delivery_terms),
        ("3. FREIGHT", po.get_freight_terms_display()),
        ("4. PACKAGING / IDENT", po.packaging_ident_terms),
        ("5. INSPECTION REPORT", po.inspection_report_terms),
        ("6. PACKING", po.packing_terms),
    ]
    for label, value in terms_rows:
        ws.cell(row=data_row, column=1, value=label)
        ws.merge_cells(start_row=data_row, start_column=2, end_row=data_row, end_column=8)
        ws.cell(row=data_row, column=2, value=value)
        data_row += 1

    widths = [8, 16, 30, 14, 10, 10, 10, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    for row_cells in ws.iter_rows(min_row=1, max_row=data_row, min_col=1, max_col=8):
        for cell in row_cells:
            if cell.value is not None:
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = Alignment(vertical="top")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def purchase_order_to_pdf(po: PurchaseOrder) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    elements = []

    left_block_rows = [[Paragraph(f"<b>{COMPANY_NAME}</b>", STYLE_BOLD_MEDIUM)]]
    left_block_rows.extend([[Paragraph(line, STYLE_NORMAL_SMALL)] for line in COMPANY_BILL_TO_LINES])
    left_block = Table(left_block_rows, colWidths=[106 * mm])
    left_block.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    right_meta_rows = [[Paragraph(f"<b>{label}</b>", STYLE_NORMAL_SMALL), Paragraph(value, STYLE_NORMAL_SMALL)] for label, value in _po_meta_rows(po)]
    right_block = Table(right_meta_rows, colWidths=[43 * mm, 39 * mm])
    right_block.setStyle(
        TableStyle(
            [
                # Outer border is handled by top_table to avoid double-line misalignment.
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    top_table = Table([[left_block, right_block]], colWidths=[108 * mm, 82 * mm])
    top_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, colors.black),
                ("LINEAFTER", (0, 0), (0, -1), 0.8, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    vendor_lines = [
        f"<b>Vendor's Name & Address</b>",
        po.vendor.name,
        po.vendor.address_line1,
        po.vendor.address_line2,
        f"{po.vendor.city}, {po.vendor.state}",
        f"State Code - {po.vendor.gst_number[:2] if po.vendor.gst_number else '-'}",
        f"GSTIN/UIN - {po.vendor.gst_number}",
    ]
    vendor_text = "<br/>".join([line for line in vendor_lines if line and line.strip()])
    vendor_table = Table([[Paragraph(vendor_text, STYLE_NORMAL_SMALL)]], colWidths=[190 * mm])
    vendor_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    item_rows = [[
        Paragraph("Sr.No", STYLE_TABLE_HEADER),
        Paragraph("Item Code", STYLE_TABLE_HEADER),
        Paragraph("Item Name", STYLE_TABLE_HEADER),
        Paragraph("HSN / SAC No", STYLE_TABLE_HEADER),
        Paragraph("Qty.", STYLE_TABLE_HEADER),
        Paragraph("UOM", STYLE_TABLE_HEADER),
        Paragraph("Rate", STYLE_TABLE_HEADER),
        Paragraph("Amount", STYLE_TABLE_HEADER),
    ]]
    total_amount = Decimal("0")
    for index, item in enumerate(po.items.select_related("material").all(), start=1):
        rate = item.unit_rate if item.unit_rate > 0 else item.material.cost_per_unit
        amount = item.line_amount if item.line_amount > 0 else (item.quantity * rate)
        total_amount += amount
        item_rows.append(
            [
                _pdf_cell(index, align="center"),
                _pdf_cell(item.material.rm_id or item.material.code or f"RM-{item.material_id}", align="center"),
                _pdf_cell(item.material.name, align="left"),
                _pdf_cell(item.material.code or "-", align="center"),
                _pdf_cell(_format_qty(item.quantity), align="center"),
                _pdf_cell(item.unit, align="center"),
                _pdf_cell(_format_money(rate), align="center"),
                _pdf_cell(_format_money(amount), align="center"),
            ]
        )

    item_rows.append(
        [
            "",
            "",
            "",
            "",
            "",
            _pdf_cell("Total", align="center", bold=True),
            "",
            _pdf_cell(_format_money(total_amount), align="center", bold=True),
        ]
    )

    item_table = Table(
        item_rows,
        colWidths=[12 * mm, 24 * mm, 55 * mm, 22 * mm, 20 * mm, 16 * mm, 18 * mm, 23 * mm],
    )
    item_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e6e6e6")),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (3, 1), (-1, -1), "CENTER"),
                ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("SPAN", (0, -1), (4, -1)),
                ("ALIGN", (0, -1), (-1, -1), "CENTER"),
            ]
        )
    )

    terms_rows = [
        [Paragraph("<b>TERMS & CONDITIONS</b>", STYLE_NORMAL_SMALL), Paragraph(f"<b>GST {po.get_freight_terms_display().upper()}</b>", STYLE_NORMAL_SMALL)],
        ["1. PAYMENTS", f"PDC against {po.payment_pdc_days} Days"],
        ["2. DELIVERY", po.delivery_terms],
        ["3. FREIGHT", po.get_freight_terms_display()],
        ["4. PACKAGING / IDENT", po.packaging_ident_terms],
        ["5. INSPECTION REPORT", po.inspection_report_terms],
        ["6. PACKING", po.packing_terms],
    ]
    terms_table = Table(terms_rows, colWidths=[68 * mm, 122 * mm])
    terms_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    prepared_sig = _load_signature_image(
        _resolve_signature_path(PREPARED_SIGNATURE_CANDIDATES),
        max_width_mm=56,
        max_height_mm=20,
    )
    authorised_sig = _load_signature_image(
        _resolve_signature_path(AUTHORISED_SIGNATURE_CANDIDATES),
        max_width_mm=56,
        max_height_mm=20,
    )

    left_sign_block = Table(
        [
            [prepared_sig],
            [Paragraph("Prepared /Checked by", STYLE_NORMAL_SMALL)],
        ],
        colWidths=[95 * mm],
    )
    left_sign_block.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    right_sign_block = Table(
        [
            [Paragraph("<b>For Busta Mobility</b>", STYLE_BOLD_SMALL)],
            [authorised_sig],
            [Paragraph("Authorised Signatory", STYLE_NORMAL_SMALL)],
        ],
        colWidths=[95 * mm],
    )
    right_sign_block.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    signature_table = Table([[left_sign_block, right_sign_block]], colWidths=[95 * mm, 95 * mm])
    signature_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    footer_table = Table(
        [
            [Paragraph(f"<b>{JURISDICTION_TEXT}</b>", STYLE_NORMAL_SMALL)],
            [Paragraph(COMPUTER_GENERATED_TEXT, STYLE_NORMAL_SMALL)],
        ],
        colWidths=[190 * mm],
    )
    footer_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    title_table = Table([[Paragraph("<b>PURCHASE ORDER</b>", STYLE_BOLD_MEDIUM)]], colWidths=[190 * mm])
    title_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    elements.extend(
        [
            title_table,
            top_table,
            vendor_table,
            item_table,
            terms_table,
            signature_table,
            footer_table,
        ]
    )

    doc.build(elements)
    return buffer.getvalue()
