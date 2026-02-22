from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


EXPORT_DIR = Path(__file__).resolve().parent / "data" / "exports"


def generate_purchase_order_excel(po: dict) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    file_path = EXPORT_DIR / f"purchase_order_{po['id']}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    ws["A1"] = f"Purchase Order #{po['id']}"
    ws["A1"].font = Font(size=16, bold=True)

    ws["A3"] = "Vendor"
    ws["B3"] = po["vendor_name"]
    ws["A4"] = "Order Date"
    ws["B4"] = po["order_date"]
    ws["A5"] = "GSTIN"
    ws["B5"] = po["gst_number"]

    address2 = po.get("address_line2") or ""
    full_address = f"{po['address_line1']}, {address2}, {po['city']}, {po['state']} - {po['pincode']}"
    ws["A6"] = "Address"
    ws["B6"] = full_address.replace(", ,", ",")
    ws["B6"].alignment = Alignment(wrap_text=True)

    ws.append([])
    ws.append(["S.No", "Material", "Code", "Quantity", "Unit"])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for idx, item in enumerate(po["items"], start=1):
        ws.append(
            [
                idx,
                item["material_name"],
                item["code"],
                item["quantity"],
                item["unit"],
            ]
        )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10

    wb.save(file_path)
    return file_path
